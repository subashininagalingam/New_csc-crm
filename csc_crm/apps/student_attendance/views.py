from itertools import count
from multiprocessing import context
# from tkinter.font import Font
from urllib import request, response
from csc_crm.apps.student_attendance.forms import BatchForm


from rest_framework import viewsets
from django.shortcuts import render
from csc_crm.apps.admissions.models import Enrollment
from django.utils import timezone
from rest_framework import generics
from rest_framework import status

from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.utils.timezone import localdate
from rest_framework import filters

from django.db import transaction
from rest_framework import status
from rest_framework.views import APIView

from .filters import AttendanceFilter

from django.http import HttpResponse
import csv

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import SimpleDocTemplate, Table, Spacer, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet

from django.http import JsonResponse
from django.db.models import Count
from django.utils import timezone

from django.http import JsonResponse
from django.db.models import Q
from django.utils import timezone
from datetime import timedelta

from django.http import JsonResponse
from csc_crm.apps.admissions.models import Course


from csc_crm.apps.admissions.models import Course



from .models import (
    Trainer,
    Batch,
    Attendance,
    SyllabusLog
)

from .serializers import (
    # TrainerSerializer,
    BatchSerializer,
    AttendanceSerializer,
    SyllabusLogSerializer
)


# class TrainerViewSet(viewsets.ModelViewSet):

#     queryset = Trainer.objects.all()

#     serializer_class = TrainerSerializer


class BatchViewSet(viewsets.ModelViewSet):

    queryset = Batch.objects.all()

    serializer_class = BatchSerializer

    filter_backends = [filters.SearchFilter]

    search_fields = [
        'batch_name',
        'timing',
        'trainer__trainer_name',
        'course__course_name'
    ]

    def create(self, validated_data):
        print("VALIDATED DATA:", validated_data)
        return super().create(validated_data)

    def create(self, request, *args, **kwargs):
        print(request.data)
        return super().create(request, *args, **kwargs)

def get_batches_by_course(request):
    course_id = request.GET.get('course_id')

    batches = Batch.objects.filter(course_id=course_id)

    data = [
        {
        "id": b.id,
        "batch_name": b.batch_name
        }
        for b in batches
        if b.student_count < 30
    ]

    return JsonResponse(data, safe=False)

class AttendanceViewSet(viewsets.ModelViewSet):

    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer

    filter_backends = [filters.SearchFilter]

    search_fields = [
        'enrollment__student_first_name',
        'enrollment__student_last_name',
        'enrollment__student_id',
    ]

    def create(self, request, *args, **kwargs):

        enrollment = request.data.get('enrollment')
        batch = request.data.get('batch')
        attendance_status = request.data.get('status')

        attendance_date = timezone.now().date()

        batch_obj = Batch.objects.get(id=batch)

        attendance, created = Attendance.objects.update_or_create(

            enrollment_id=enrollment,
            batch_id=batch,
            attendance_date=attendance_date,

            defaults={
                'status': attendance_status,
                'trainer': batch_obj.trainer
            }
        )

        serializer = self.get_serializer(attendance)

        return Response(
            serializer.data,
            status=status.HTTP_200_OK
        )
    
class AttendanceSubmitAPIView(APIView):

    @transaction.atomic
    def post(self, request):

        batch_id = request.data.get("batch")

        attendance_list = request.data.get(
            "attendance",
            []
        )

        syllabus_data = request.data.get(
            "syllabus_log",
            {}
        )

        try:

            batch = Batch.objects.get(
                id=batch_id
            )

        except Batch.DoesNotExist:

            return Response(
                {
                    "status": False,
                    "message": "Batch not found"
                },
                status=status.HTTP_404_NOT_FOUND
            )

        for item in attendance_list:

            Attendance.objects.update_or_create(

                enrollment_id=item["enrollment"],

                batch=batch,

                attendance_date=timezone.now().date(),

                defaults={
                    "status": item["status"],
                    "remarks": item.get(
                        "remarks",
                        ""
                    ),
                    "trainer": batch.staff
                }
            )

        SyllabusLog.objects.create(

            batch=batch,

            trainer=batch.staff,

            topic_covered=syllabus_data[
                "topic_covered"
            ],

            duration=syllabus_data[
                "duration"
            ],

            next_topic=syllabus_data.get(
                "next_topic"
            ),

            trainer_notes=syllabus_data.get(
                "trainer_notes"
            )
        )

        return Response(
            {
                "status": True,
                "message":
                "Attendance and syllabus log saved successfully"
            },
            status=status.HTTP_200_OK
        )
    
class SyllabusLogViewSet(viewsets.ModelViewSet):
    queryset = SyllabusLog.objects.all().order_by('-date')
    serializer_class = SyllabusLogSerializer



def get_course_duration(request, course_id):

    try:
        course = Course.objects.get(id=course_id)

        return JsonResponse({
            "duration": course.duration
        })

    except Course.DoesNotExist:

        return JsonResponse(
            {"error": "Course not found"},
            status=404
        )


def dashboard_api(request):

    search = request.GET.get("search", "")
    today = timezone.now().date()

    batches = Batch.objects.all()

    if search:
        batches = batches.filter(
        Q(batch_name__icontains=search) |
        Q(course__course_name__icontains=search) |
        Q(trainer__first_name__icontains=search) |
        Q(trainer__last_name__icontains=search) |
        Q(timing__icontains=search)
    )

    enrollments = Enrollment.objects.filter(batch__in=batches)

    attendance_qs = Attendance.objects.filter(
        enrollment__in=enrollments,
        attendance_date=today
    )

    batch_details = []

    for batch in batches:
        
        if Enrollment.objects.filter(batch=batch).exists():
            
            batch_details.append({
            "course": batch.course.course_name,
            "batch": batch.batch_name,
            "trainer": (f"{batch.trainer.first_name} {batch.trainer.last_name}"if batch.trainer else "Not Assigned"),
            "timing": f"{batch.start_time.strftime('%I:%M %p')} - {batch.end_time.strftime('%I:%M %p')}",
            "session": batch.timing,
        })

    return JsonResponse({
        "total": enrollments.count(),
        "present": attendance_qs.filter(status="Present").count(),
        "absent": attendance_qs.filter(status="Absent").count(),
        "late": attendance_qs.filter(status="Late").count(),
        "batches": batch_details,
    })

def dashboard(request):

    search = request.GET.get("search", "")
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)

    batches = Batch.objects.all()

    if search:
        batches = batches.filter(
            Q(batch_name__icontains=search) |
            Q(course__course_name__icontains=search) |
            Q(trainer__first_name__icontains=search) |
            Q(trainer__last_name__icontains=search) |
            Q(timing__icontains=search)
        )

    enrollments = Enrollment.objects.filter(
        batch__in=batches
    )

    total= enrollments.count()

    attendance_qs = Attendance.objects.filter(
        enrollment__in=enrollments,
        attendance_date=today
    )

  

    today = timezone.now().date()

    # Current Month
    current_month_students = Enrollment.objects.filter(
    start_date__year=today.year,
    start_date__month=today.month
    ).count()

    # Previous Month
    if today.month == 1:
        prev_month = 12
        prev_year = today.year - 1
    else:
        prev_month = today.month - 1
        prev_year = today.year

    previous_month_students = Enrollment.objects.filter(
    start_date__year=prev_year,
    start_date__month=prev_month
    ).count()

    # Percentage
    if previous_month_students > 0:
        total_percentage = round(
            ((current_month_students - previous_month_students) / previous_month_students) * 100, 2
        )
    else:
        total_percentage = 100 if current_month_students > 0 else 0

    present = attendance_qs.filter(status="Present").count()
    absent = attendance_qs.filter(status="Absent").count()
    late = attendance_qs.filter(status="Late").count()

    attendance_marked = attendance_qs.count()

    last_attendance_date = Attendance.objects.filter(
    enrollment__in=enrollments,
    attendance_date__lt=today
    ).order_by('-attendance_date').values_list(
        'attendance_date',
       flat=True
    ).first()

    last_present = 0
    last_absent = 0
    last_late = 0

    if last_attendance_date:
        last_qs = Attendance.objects.filter(
        enrollment__in=enrollments,
        attendance_date=last_attendance_date
        )

        last_present = last_qs.filter(status="Present").count()
        last_absent = last_qs.filter(status="Absent").count()
        last_late = last_qs.filter(status="Late").count()
    percentage = round((present / total) * 100, 2) if total else 0

    #present_change = 0
    #absent_change = 0
    #late_change = 0
    
    present_percentage = round((present / total) * 100, 2) if total else 0
    absent_percentage = round((absent / total) * 100, 2) if total else 0
    late_percentage = round((late / total) * 100, 2) if total else 0
    
    last_total = last_present + last_absent + last_late

    last_present_percentage = round(
        (last_present / last_total) * 100, 2
    ) if last_total else 0

    last_absent_percentage = round(
        (last_absent / last_total) * 100, 2
    ) if last_total else 0

    last_late_percentage = round(
        (last_late / last_total) * 100, 2
    ) if last_total else 0

    present_change = round(
        present_percentage - last_present_percentage,
        2
    )

    absent_change = round(
        absent_percentage - last_absent_percentage,
        2
    )

    late_change = round(
        late_percentage - last_late_percentage,
        2
    )

    #if last_present > 0:
     #   present_change = round(
      #  ((present - last_present) / last_present) * 100, 2
    #)

    #if last_absent > 0:
      #  absent_change = round(
      #  ((absent - last_absent) / last_absent) * 100, 2
    #)

    #if last_late > 0:
    #    late_change = round(
     #   ((late - last_late) / last_late) * 100, 2
  #  )
  
    
    batch_details = []

    for batch in batches:

        if Enrollment.objects.filter(batch=batch).exists():

            batch_details.append({
                "course": batch.course.course_name,
                "batch": batch.batch_name,
                "trainer": (f"{batch.trainer.first_name} {batch.trainer.last_name}"if batch.trainer else "Not Assigned"),
                "timing": f"{batch.start_time.strftime('%I:%M %p')} - {batch.end_time.strftime('%I:%M %p')}"
            })
        
                
                

    context = {
        "present": present,
        "absent": absent,
        "late": late,
        "total": total,
        "percentage": percentage,
        "attendance_marked": attendance_marked,
        "present_change": present_change,
        "absent_change": absent_change,
        "late_change": late_change,
        "total_percentage":total_percentage,
        "context_batches": batch_details,
        "present_percentage": present_percentage,
        "absent_percentage": absent_percentage,
        "late_percentage": late_percentage,
        "last_attendance_date": last_attendance_date,
        
    }

    return render(
        request,
        'attendance/dashboard.html',
        context
    )


def batches_page(request):

    form = BatchForm()

    context = {
        'form': form
    }

    return render(
        request,
        'attendance/batches.html',
        context
    )


def mark_attendance_page(request, batch_id):

    batch = Batch.objects.get(
        id=batch_id
    )

    enrollments = Enrollment.objects.filter(
        batch=batch,
        admission__course_name=batch.course,
    )

    attendance_records = Attendance.objects.filter(
        batch=batch,
        attendance_date=timezone.now().date()
    )

    syllabus_log = SyllabusLog.objects.filter(
    batch=batch,
    date=timezone.now().date()
    ).first()

    duration = syllabus_log.duration if syllabus_log else 0

    attendance_map = {
        att.enrollment_id: att.status
        for att in attendance_records
    }

    remarks_map = {
        att.enrollment_id: att.remarks
        for att in attendance_records
    }

    context = {
        'batch': batch,
        'enrollments': enrollments,
        'attendance_map': attendance_map,
        'remarks_map': remarks_map,
        'syllabus_log': syllabus_log,
        "duration_hours": duration // 60,
        "duration_minutes": duration % 60,
    }

    return render(
        request,
        'attendance/mark_attendance.html',
        context
    )



@api_view(['GET'])
def today_attendance_summary(request, batch_id):

    today = timezone.now().date()

    qs = Attendance.objects.filter(
        batch_id=int(batch_id),  
        attendance_date=today
    )

    data = {
        "present": qs.filter(status="Present").count(),
        "absent": qs.filter(status="Absent").count(),
        "late": qs.filter(status="Late").count(),
        "total": qs.count()
    }

    return Response(data)


@api_view(['POST'])
@transaction.atomic
def bulk_attendance(request):

    print("REQUEST DATA:")
    print(request.data)

    batch_id = request.data.get('batch')

    attendance_list = request.data.get(
        'attendance',
        []
    )

    syllabus_data = request.data.get(
        'syllabus_log',
        {}
    )

    if not syllabus_data.get('topic_covered'):
        return Response(
            {
                'status': False,
                'message': 'Topic covered is required'
            },
            status=400
        )

    duration = syllabus_data.get('duration')

    if not duration:
        return Response(
            {
                'status': False,
                'message': 'Duration is required'
            },
            status=400
        )

    try:

        batch_obj = Batch.objects.get(
            id=batch_id
        )

    except Batch.DoesNotExist:

        return Response(
            {
                'status': False,
                'message': 'Batch not found'
            },
            status=404
        )

    attendance_date = timezone.now().date()

    for item in attendance_list:

        Attendance.objects.update_or_create(

            enrollment_id=item['enrollment'],

            batch_id=batch_id,

            attendance_date=attendance_date,

            defaults={
                'status': item['status'],
                'remarks': item.get(
                    'remarks',
                    ''
                ),
                'trainer': batch_obj.trainer
            }
        )

    SyllabusLog.objects.update_or_create(

    batch=batch_obj,

    date=attendance_date,

    defaults={
        'trainer': batch_obj.trainer,
        'topic_covered': syllabus_data.get(
            'topic_covered'
        ),
        'duration': syllabus_data.get(
            'duration'
        ),
        'next_topic': syllabus_data.get(
            'next_topic',
            ''
        ),
        'trainer_notes': syllabus_data.get(
            'trainer_notes',
            ''
        )
    }
)

    return Response({
        'status': True,
        'message':
        'Attendance and syllabus log saved successfully'
    })

def attendance_history_page(request):

    records = Attendance.objects.select_related(
        'enrollment',
        'batch'
    ).order_by('-attendance_date')
    
    courses = Course.objects.all()

    attendance_filter = AttendanceFilter(
        request.GET,
        queryset=records
    )

    print("GET:", request.GET)
    print("COUNT:", attendance_filter.qs.count())

    context = {
        "filter": attendance_filter,
        "records": attendance_filter.qs,
        "courses": courses
    }

    return render(
        request,
        'attendance/attendance_history.html',
        context
    )


def attendance_export(request):

    records = Attendance.objects.select_related(
    'enrollment',
    'batch',
    'trainer'
).order_by('-attendance_date')

    attendance_filter = AttendanceFilter(request.GET, queryset=records)
    qs = attendance_filter.qs

    export_format = request.GET.get("format")

    # ================= EXCEL =================
    if export_format == "excel":

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        headers = ['Date', 'Student', 'Course', 'Batch', 'Status', 'Trainer']
        ws.append(headers)

        # 🪔 Temple Gold Header Style
        header_fill = PatternFill(
            start_color="FFC000",
            end_color="FFC000",
            fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in qs:
            student_name = f"{r.enrollment.student.first_name} {r.enrollment.student.last_name}"
            course_name = r.enrollment.course.course_name
            batch_name = r.batch.batch_name
            trainer_name = getattr(r.trainer, "name", str(r.trainer))

            ws.append([
                str(r.attendance_date),
                student_name,
                course_name,
                batch_name,
                str(r.status),
                trainer_name
            ])

        # 📏 Column Widths (IMPORTANT: moved OUTSIDE loop)
        column_widths = {
            'A': 20,
            'B': 25,
            'C': 20,
            'D': 18,
            'E': 15,
            'F': 20,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'

        wb.save(response)
        return response

    # ================= PDF =================
    elif export_format == "pdf":

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)

        styles = getSampleStyleSheet()
        title = Paragraph("🪔 Attendance Report", styles['Title'])

        data = [['Date', 'Student', 'Course', 'Batch', 'Status', 'Trainer']]

        for r in qs:
            student_name = f"{r.enrollment.student.first_name} {r.enrollment.student.last_name}"
            course_name = r.enrollment.course.course_name
            batch_name = r.batch.batch_name
            trainer_name = getattr(r.trainer, "name", str(r.trainer))

            data.append([
                str(r.attendance_date),
                student_name,
                course_name,
                batch_name,
                str(r.status),
                trainer_name
            ])

        table = Table(data)

        table.setStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.gold),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),

            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),

            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ])

        elements = [
            title,
            Spacer(1, 12),
            table
        ]

        doc.build(elements)
        return response

    return HttpResponse("Invalid format", status=400)


def student_attendance_summary(request, student_id):

    records = Attendance.objects.filter(
        enrollment__admission__student_id=student_id
    ).order_by("-attendance_date")

    present = records.filter(status="Present").count()
    absent = records.filter(status="Absent").count()
    late = records.filter(status="Late").count()

    total = records.count()

    percentage = round(((present + late) / total) * 100, 2) if total else 0

    student = records.first().enrollment.student if records.exists() else None

    timeline = []

    first_record = records.first()

    for r in records:
        timeline.append({
            "date": r.attendance_date.strftime("%d-%m-%Y"),
            "status": r.status
        })

    return JsonResponse({
        "student_name": (
            f"{student.first_name} {student.last_name}"
            if student else ""
        ),
        "student_id": f"STU{student.id}",
        "course": first_record.enrollment.course.course_name if first_record else "",
        "batch": first_record.batch.batch_name if first_record else "",
        "timing": first_record.batch.timing if first_record else "",
        "photo_url": student.photo.url if student and student.photo else None,
        "present": present,
        "absent": absent,
        "late": late,
        "percentage": percentage,
        "timeline": timeline
    })


         # absent tracker

from .services import get_absent_tracker_data
from .models import AbsentTracker


def absent_tracker(request):

    absent_students = get_absent_tracker_data()

    absent_students = [
        student
        for student in absent_students
        if student['total_absences'] > 0
    ]

    absent_students = sorted(
        absent_students,
        key=lambda x: (
            x['consecutive_absences'],
            x['total_absences']
        ),
        reverse=True
    )
    
    courses = Course.objects.all()
    batches = Batch.objects.all()

    return render(
        request,
        'attendance/absent_tracker.html',
        {
            'absent_students': absent_students,
            'courses': courses,
            'batches': batches,
        }
    )
    
from django.http import JsonResponse
from django.utils import timezone
from .models import AbsentTracker

def mark_notification_sent(request, enrollment_id):
    print("MARK NOTIFICATION CALLED", enrollment_id)

    tracker, created = AbsentTracker.objects.get_or_create(
        enrollment_id=enrollment_id
    )

    tracker.notification_sent = True
    tracker.notification_status = "Dispatched"
    tracker.last_notified_at = timezone.now()
    tracker.save()

    return JsonResponse({
        "status": "success"
    })
    
def get_admin_notes(request, tracker_id):

    tracker = AbsentTracker.objects.filter(
        id=tracker_id
    ).first()

    return JsonResponse({

        "notes":
        tracker.admin_notes
        if tracker and tracker.admin_notes
        else ""

    })
    
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import json

@require_POST
def save_admin_notes(request):

    data = json.loads(request.body)
    
    print("REQUEST DATA =", data)
    print("TRACKER ID =", data.get("tracker_id"))

    tracker_id = data.get("tracker_id")

    if not tracker_id or tracker_id == "None":
        return JsonResponse({
        "status": "error",
        "message": "Tracker ID not found"
    }, status=400)

    tracker = AbsentTracker.objects.filter(
    id=int(tracker_id)
    ).first()

    if not tracker:
        return JsonResponse({
        "status": "error",
        "message": "Tracker not found"
    })

    notes = data.get("notes", "")

    tracker.admin_notes = notes
    tracker.save()

    return JsonResponse({
        "status": "success"
    })

from django.shortcuts import render, get_object_or_404
from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages

def low_attendance_alerts(request):

    low_attendance_students = (
        get_low_attendance_data()
    )

    # --- Student name search (new) ---
    name_filter = request.GET.get("name")

    if name_filter:
        low_attendance_students = [
            student for student in low_attendance_students
            if name_filter.lower() in (
                f"{student['student'].first_name} {student['student'].last_name}".lower()
            )
        ]

    # --- Course filter ---
    course_filter = request.GET.get("course")

    if course_filter:
        low_attendance_students = [
            student for student in low_attendance_students
            # FIX: student["course"] is a Course object, not a string,
            # so we compare against .course_name instead of calling
            # .lower() directly on the object (this was a bug before).
            if student["course"].course_name.lower() == course_filter.lower()
        ]

    # --- Batch filter ---
    batch_filter = request.GET.get("batch")

    if batch_filter:
        low_attendance_students = [
            student for student in low_attendance_students
            # FIX: compare by batch id (matches the <option value="{{ batch.id }}">
            # in the template) instead of calling .lower() on the Batch object.
            if str(student["batch"].id) == str(batch_filter)
        ]

    # --- Attendance % filter ---
    attendance_filter = request.GET.get("attendance")

    if attendance_filter:
        low_attendance_students = [
            student
            for student in low_attendance_students
            if student["attendance_percentage"] == float(attendance_filter)
        ]

    critical_students = [
        student
        for student in low_attendance_students
        if student["alert_level"] == "Critical"
    ]

    warning_students = [
        student
        for student in low_attendance_students
        if student["alert_level"] == "Warning"
    ]

    # --- Stats for the top cards (Critical / Warning / Total / Average) ---
    total_students = len(low_attendance_students)

    overall_average = (
        round(
            sum(s["attendance_percentage"] for s in low_attendance_students)
            / total_students,
            1
        )
        if total_students > 0
        else 0
    )

    courses = Course.objects.all()

    batches = Batch.objects.all()

    return render(

        request,

        'attendance/low_attendance.html',

        {
            'low_attendance_students': low_attendance_students,

            "critical_students": critical_students,
            "warning_students": warning_students,

            "total_students": total_students,
            "overall_average": overall_average,

            'courses': courses,

            'batches': batches,

            "filters": {
                "name": name_filter or "",
                "course": course_filter or "",
                "batch": batch_filter or "",
                "attendance": attendance_filter or "",
            },
        }

    )


from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill

def low_attendance_export(request):

    alert_type = request.GET.get("type")

    students = get_low_attendance_data()

    if alert_type == "critical":
        students = [
            s for s in students
            if s["alert_level"] == "Critical"
        ]
        file_name = "critical_alerts.xlsx"

    else:
        students = [
            s for s in students
            if s["alert_level"] == "Warning"
        ]
        file_name = "warning_alerts.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Low Attendance"

    headers = [
        "Student Name",
        "Course",
        "Batch",
        "Attendance %",
        "Consecutive Absences",
        "Total Absences"
    ]

    ws.append(headers)

    header_fill = PatternFill(
        start_color="1E40AF",
        end_color="1E40AF",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )
        cell.fill = header_fill

    for student in students:

        ws.append([
            f"{student['student'].first_name} {student['student'].last_name}",
            student['course'].course_name,
            student['batch'].batch_name,
            student['attendance_percentage'],
            student['consecutive_absences'],
            student['total_absences']
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = (
        f'attachment; filename="{file_name}"'
    )

    wb.save(response)

    return response


def send_low_attendance_email(request, enrollment_id):

    enrollment = get_object_or_404(
        Enrollment,
        id=enrollment_id
    )

    student = enrollment.student

    attendance_records = Attendance.objects.filter(
        enrollment=enrollment
    )

    total_working_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()

    present_count = attendance_records.filter(
        status='Present'
    ).count()

    attendance_percentage = (
        round(
            (present_count / total_working_days) * 100,
            1
        )
        if total_working_days > 0
        else 100
    )

    if attendance_percentage < 60:

        subject = "Critical Attendance Alert"

        message = f"""
Dear {student.first_name},

Your attendance percentage is critically low.

Attendance Percentage: {attendance_percentage}%

Immediate action is required.

Please contact your trainer.

Regards,
CSC Computer Education
"""

    else:

        subject = "Low Attendance Warning"

        message = f"""
Dear {student.first_name},

Your attendance percentage is below the required level.

Attendance Percentage: {attendance_percentage}%

Please attend classes regularly.

Regards,
CSC Computer Education
"""

    send_mail(

        subject,

        message,

        settings.DEFAULT_FROM_EMAIL,

        [student.email],

        fail_silently=False

    )

    messages.success(
        request,
        f"Email sent to {student.email}"
    )

    return JsonResponse({
        "message": "📧 Email sent successfully"
    })


def send_sms_notification(request, enrollment_id):

    enrollment = get_object_or_404(
        Enrollment,
        id=enrollment_id
    )

    student = enrollment.student

    attendance_records = Attendance.objects.filter(
        enrollment=enrollment
    )

    total_working_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()

    present_count = attendance_records.filter(
        status='Present'
    ).count()

    attendance_percentage = (
        round(
            (present_count / total_working_days) * 100,
            1
        )
        if total_working_days > 0
        else 100
    )

    if attendance_percentage < 60:

        sms_message = f"""
CSC ALERT

Critical Attendance Alert

Student:
{student.first_name}

Attendance:
{attendance_percentage}%

Immediate action required.
"""

    else:

        sms_message = f"""
CSC ALERT

Low Attendance Warning

Student:
{student.first_name}

Attendance:
{attendance_percentage}%

Please attend classes regularly.
"""

    print(sms_message)

    messages.success(
        request,
        f"SMS sent to {student.phone_no}"
    )

    return JsonResponse({
        "message": "📱 SMS sent successfully"
    })


@require_POST
def send_bulk_notification(request):
    """
    NEW endpoint used by the redesigned page.

    Sends SMS/Email only to the students whose checkboxes are
    ticked inside a given alert section ("Send SMS" / "Send Email"
    buttons in the Critical / Warning section headers).

    Expects a JSON body like:
        { "type": "email" | "sms", "enrollment_ids": [12, 14, 19] }

    Add this to your urls.py:
        path('api/send-bulk-notification/', views.send_bulk_notification,
             name='send_bulk_notification'),
    """

    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({"message": "Invalid request data"}, status=400)

    notification_type = payload.get("type")
    enrollment_ids = payload.get("enrollment_ids", [])

    if notification_type not in ("sms", "email"):
        return JsonResponse({"message": "Invalid notification type"}, status=400)

    if not enrollment_ids:
        return JsonResponse({"message": "No students selected"}, status=400)

    total_working_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()

    sent_count = 0

    for enrollment_id in enrollment_ids:

        enrollment = Enrollment.objects.filter(id=enrollment_id).first()

        if not enrollment:
            continue

        student = enrollment.student

        present_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Present'
        ).count()

        attendance_percentage = (
            round(
                (present_count / total_working_days) * 100,
                1
            )
            if total_working_days > 0
            else 100
        )

        is_critical = attendance_percentage < 60

        if notification_type == "email":

            if not student.email:
                continue

            subject = (
                "Critical Attendance Alert"
                if is_critical else "Low Attendance Warning"
            )

            message = f"""
Dear {student.first_name},

Your attendance percentage is {"critically low" if is_critical else "below the required level"}.

Attendance Percentage: {attendance_percentage}%

{"Immediate action is required." if is_critical else "Please attend classes regularly."}

Regards,
CSC Computer Education
"""

            send_mail(

                subject,

                message,

                settings.DEFAULT_FROM_EMAIL,

                [student.email],

                fail_silently=False

            )

            sent_count += 1

        else:

            sms_message = f"""
CSC ALERT

{"Critical Attendance Alert" if is_critical else "Low Attendance Warning"}

Student:
{student.first_name}

Attendance:
{attendance_percentage}%
"""

            print(sms_message)

            sent_count += 1

    label = "Emails" if notification_type == "email" else "SMS"

    return JsonResponse({
        "message": f"✅ {label} sent to {sent_count} student(s)"
    })


def send_email_all(request):

    low_attendance_students = (
        get_low_attendance_data()
    )

    for student_data in low_attendance_students:

        student = student_data["student"]

        attendance_percentage = (
            student_data["attendance_percentage"]
        )

        if attendance_percentage < 60:

            subject = (
                "Critical Attendance Alert"
            )

            message = f"""
Dear {student.first_name},

Your attendance percentage is critically low.

Attendance:
{attendance_percentage}%

Immediate action required.

Regards,
CSC Computer Education
"""

        else:

            subject = (
                "Low Attendance Warning"
            )

            message = f"""
Dear {student.first_name},

Your attendance percentage is below the required level.

Attendance:
{attendance_percentage}%

Please improve attendance.

Regards,
CSC Computer Education
"""

        if student.email:

            send_mail(

                subject,

                message,

                settings.DEFAULT_FROM_EMAIL,

                [student.email],

                fail_silently=False

            )

    return JsonResponse({

        "message":
        "All Emails sent successfully"

    })


def send_sms_all(request):

    low_attendance_students = (
        get_low_attendance_data()
    )

    for student_data in low_attendance_students:

        student = student_data["student"]

        attendance_percentage = (
            student_data["attendance_percentage"]
        )

        if attendance_percentage < 60:

            sms_message = f"""
CSC ALERT

Critical Attendance Alert

Student:
{student.first_name}

Attendance:
{attendance_percentage}%
"""

        else:

            sms_message = f"""
CSC ALERT

Low Attendance Warning

Student:
{student.first_name}

Attendance:
{attendance_percentage}%
"""

        print(sms_message)

    messages.success(
        request,
        "SMS notifications sent."
    )

    return JsonResponse({
        "message": " All SMS notifications sent"
    })


def send_monthly_report(request):

    low_attendance_students = (
        get_low_attendance_data()
    )

    report_lines = []

    report_lines.append(
        "Monthly Low Attendance Report\n"
    )

    for student_data in low_attendance_students:

        report_lines.append(

            f"""
Student:
{student_data['student'].first_name}
{student_data['student'].last_name}

Course:
{student_data['course'].course_name}

Batch:
{student_data['batch'].batch_name}

Attendance:
{student_data['attendance_percentage']}%

Total Absences:
{student_data['total_absences']}
"""
        )

    report_content = "\n".join(
        report_lines
    )

    send_mail(

        "Monthly Attendance Report",

        report_content,

        settings.DEFAULT_FROM_EMAIL,

        [settings.DEFAULT_FROM_EMAIL],

        fail_silently=False

    )

    messages.success(
        request,
        "Monthly report sent."
    )

    return JsonResponse({
        "message": "📊 Monthly report sent successfully"
    })
    
    
    #Reports 
    
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Count

from csc_crm.apps.admissions.models import Enrollment, Course
from .models import Attendance, Batch
from .services import get_low_attendance_data
from django.db.models import Q

def get_report_students():

    report_students = []

    enrollments = Enrollment.objects.select_related(
        'admission__student',
        'admission__course_name',
        'batch'
    )

    total_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()

    for enrollment in enrollments:

        present_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Present'
        ).count()
        absent_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Absent'
        ).count()

        late_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Late'
        ).count()

        total_days = (
            present_count +
            absent_count +
            late_count
        )

        effective_present = (
            present_count +
            late_count
        )

        attendance_rate = round(
            (effective_present / total_days) * 100,
            1
        ) if total_days > 0 else 0

        if attendance_rate >= 100:
            status = "Excellent"

        elif attendance_rate >= 75:
            status = "Good"
            
        elif attendance_rate >= 60:
            status = "Warning"

        else:
            status = "Critical"

        report_students.append({

            #"student": enrollment.student,
            "student": enrollment.admission.student,

            #"course": enrollment.course,
            "course": enrollment.admission.course_name,

            "batch": enrollment.batch,

            "present_count": present_count,

            "absent_count": absent_count,

            "late_count": late_count,

            "attendance_rate": attendance_rate,

            "status": status,

            "total_days": total_days,

        })

    return report_students


def reports(request):

    today = timezone.now().date()
    
    latest_attendance = Attendance.objects.order_by(
        '-attendance_date'
    ).first()

    latest_attendance_date = None

    if latest_attendance:
        latest_attendance_date = latest_attendance.attendance_date
    
    attendance_date = ""
    
    if latest_attendance:
        attendance_date = latest_attendance.attendance_date.strftime(
            "%d %b %Y"
        )
    #filters
    
    student_name = request.GET.get(
        "student_name"
    )
    attendance_filter = request.GET.get(
        "attendance"
    )
    
    course_filter = request.GET.get(
        "course"
    )
    
    batch_filter = request.GET.get(
        "batch"
    )
    #selected_batch = ""

    #if batch_filter:
      #  selected_batch = batch_filter.split("|")[0]
        
    status_filter = request.GET.get(
        "status"
    )
    
    batch_chart_title = "Batch-wise Attendance"
    
    if batch_filter:
       # batch_chart_title = (
          #  f"Batch-wise Attendance - "
          #  f"{batch_filter.title()}"
       # )
       selected_batch_obj = Batch.objects.filter(
           id = batch_filter
       ).first()
       
       if selected_batch_obj:
           batch_chart_title = (
               f"Batch-wise Attendance - "
               f"{selected_batch_obj.batch_name}"
               f"({selected_batch_obj.course.course_name})"
           )
    elif course_filter:
        batch_chart_title = (
            f"Batch-wise Attendance - "
            f"{course_filter.title()}"
        )
            
    
    total_students = Enrollment.objects.count()
    
    today_marked_count = Attendance.objects.filter(
        attendance_date=today
    ).count()

    pending_count = total_students - today_marked_count

    if today_marked_count == 0:
 
       attendance_status = "not_started"

    elif pending_count > 0:

        attendance_status = "in_progress"

    else:

         attendance_status = "completed"
    
    

    # Top cards

    

    present_today = Attendance.objects.filter(
        attendance_date=today,
        ).filter(
            Q(status='Present') |
            Q(status='Late')
        ).count()

    absent_today = Attendance.objects.filter(
        attendance_date=today,
        status='Absent'
    ).count()

    low_attendance = len(
        get_low_attendance_data()
    )

    # Report table

    report_students = []

    enrollments = Enrollment.objects.select_related(
        'admission__student',
        'admission__course_name',
        'batch'
    )
    
    # enrollment filter
    
    if course_filter:
        enrollments = enrollments.filter(
            admission__course_name__course_name__iexact=
            course_filter
        )
    if batch_filter:
        enrollments = enrollments.filter(
            #batch__batch_name__iexact=
           # selected_batch
            batch_id=batch_filter
        )
    if student_name:
        enrollments = enrollments.filter(
            admission__student__first_name__istartswith=
            student_name
        )
    total_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()
    
    filtered_students = []

    if status_filter or attendance_filter:

        for enrollment in enrollments:

            present_count = Attendance.objects.filter(
                enrollment=enrollment,
                status='Present'
            ).count()

            absent_count = Attendance.objects.filter(
                enrollment=enrollment,
                status='Absent'
            ).count()

            late_count = Attendance.objects.filter(
                enrollment=enrollment,
                status='Late'
            ).count()

            total_days = (
                present_count +
                absent_count +
                late_count
            )

            effective_present = (
                present_count +
                late_count
            )

            attendance_rate = round(
            (effective_present / total_days) * 100,
            1
            ) if total_days else 0
            
            if attendance_filter:
                if round(attendance_rate, 1) != round(float(attendance_filter), 1):
                    continue

            if attendance_rate == 100:
                status = "Excellent"

            elif attendance_rate >= 75:
                status = "Good"
            
            elif attendance_rate >= 60:
                status = "Warning"

            else:
                status = "Critical"

            #if status.lower() == status_filter.lower():

               # filtered_students.append(
                #    enrollment.id
               # )
            if status_filter:

                if status.lower() == status_filter.lower():

                    filtered_students.append(
                        enrollment.id
                    )

            else:

                filtered_students.append(
                    enrollment.id
                )

        enrollments = enrollments.filter(
            id__in=filtered_students
        )   
        
    

    for enrollment in enrollments:

        present_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Present'
        ).count()

        absent_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Absent'
        ).count()

        late_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Late'
        ).count()

        total_days = (
            present_count +
            absent_count +
            late_count
        )

        effective_present = (
            present_count +
            late_count
        )

        attendance_rate = round(
            (effective_present / total_days) * 100,
            1
        ) if total_days else 0

    
        if attendance_filter:
                if round(attendance_rate, 1) != round(float(attendance_filter), 1):
                    continue

        if attendance_rate == 100:
            status = "Excellent"

        elif attendance_rate >= 75:
            status = "Good"
            
        elif attendance_rate >= 60:
            status = "Warning"

        else:
            status = "Critical"

        report_students.append({

            "student":
            enrollment.admission.student,

            "course":
            enrollment.admission.course_name,

            "batch":
            enrollment.batch,

            "present_count":
            present_count,

            "absent_count":
            absent_count,

            "late_count":
            late_count,

            "attendance_rate":
            attendance_rate,

            "status":
            status,
            
            "total_days": total_days,

        })

    # Monthly Chart
    
    monthly_chart_title = "Monthly Attendance Chart"

    #monthly_present = []
    #monthly_absent = []
    #monthly_late = []
    monthly_present = [0] * 12
    monthly_absent = [0] * 12
    monthly_late = [0] * 12
    
    
    
    attendance_qs = Attendance.objects.all()
    
    if student_name:
        attendance_qs = attendance_qs.filter(
            enrollment__admission__student__first_name__istartswith=
            student_name
        )
    if status_filter or attendance_filter:

        student_ids = []

        for enrollment in enrollments:

            student_ids.append(
                enrollment.id
            )
        print("Attendance Filter =", attendance_filter)
        print("Student IDs =", student_ids)
        print("Enrollment Count =", enrollments.count())

        attendance_qs = attendance_qs.filter(
            enrollment_id__in=student_ids
        )
    
    if course_filter:
        attendance_qs = attendance_qs.filter(
            enrollment__admission__course_name__course_name__icontains=
            course_filter
        )
        
    if batch_filter:
        attendance_qs = attendance_qs.filter(
            #  enrollment__batch__batch_name__iexact=batch_filter
            enrollment__batch_id=batch_filter
        )

    for month in range(1, 13):

        monthly_present[month - 1] = attendance_qs.filter(
            attendance_date__month=month,
            attendance_date__year=today.year,
            status="Present"
        ).count()

        monthly_absent[month - 1] = attendance_qs.filter(
            attendance_date__month=month,
            attendance_date__year=today.year,
            status="Absent"
        ).count()

        monthly_late[month - 1] = attendance_qs.filter(
            attendance_date__month=month,
            attendance_date__year=today.year,
            status="Late"
        ).count()

    # Course Analytics

    course_labels = []
    course_counts = []

    courses = Course.objects.all()

    for course in courses:

        course_labels.append(
            course.course_name
        )

        course_counts.append(

            Enrollment.objects.filter(
                admission__course_name=course
            ).count()

        )

    # Batch Analytics
    filtered_enrollment_ids = [
        enrollment.id
        for enrollment in enrollments
    ]
    
    print("Batch Chart IDs =", filtered_enrollment_ids)

    batch_labels = []
    batch_counts = []
    batch_present_counts = []
    batch_performance_labels = []
    batch_performance_counts = []
    batch_present_list = []
    batch_absent_list = []
    batch_percentage_list = []
    batch_late_list = []
    

    batches = Batch.objects.all()
    
    if course_filter:
        batches = batches.filter(
            course__course_name__icontains=
            course_filter
        )
    if batch_filter:
        batches = batches.filter(
           # batch_name__icontains=
            #batch_filter
            id=batch_filter

            
        )
        

    for batch in batches:

        batch_labels.append(
            f"{batch.course.course_name} - {batch.batch_name}"
        )

        batch_counts.append(

            Enrollment.objects.filter(
                batch=batch
            ).count()

        )
        
        
        
        present_count = Attendance.objects.filter(
            enrollment_id__in=filtered_enrollment_ids,
            batch=batch,
            attendance_date=latest_attendance_date,
            status="Present"
        ).count()
        
        absent_count = Attendance.objects.filter(
            enrollment_id__in=filtered_enrollment_ids,
            batch=batch,
            attendance_date=latest_attendance_date,
            status="Absent"
        ).count()
        
        late_count = Attendance.objects.filter(
            enrollment_id__in=filtered_enrollment_ids,
            batch=batch,
            attendance_date=latest_attendance_date,
            status="Late"
        ).count()

        total_count = Attendance.objects.filter(
            enrollment_id__in=filtered_enrollment_ids,
            batch=batch,
            attendance_date=latest_attendance_date
        ).count()

       # percentage = round(
        #    (present_count / total_count) * 100,
         #   1
        #) if total_count else 0
        
        effective_present = (
            present_count +
            late_count
        )
        percentage = round(
            (effective_present / total_count) * 100,
            1
        ) if total_count else 0
        
        batch_present_list.append(
            present_count
        )
        
        batch_absent_list.append(
            absent_count
        )
        
        batch_late_list.append(
            late_count
        )
        
        batch_percentage_list.append(
            percentage
        )

        batch_present_counts.append(
            percentage
        )

        batch_performance_labels.append(
            f"{batch.course.course_name} - {batch.batch_name}"
        )

        batch_performance_counts.append(
            percentage
        )
        
        monthly_chart_title = "Monthly Attendance Chart"
        
        if course_filter:
            monthly_chart_title = (
                f"Monthly Attendance Chart - "
                f"{course_filter.title()}"
            )
        print("Batch Labels =", batch_labels)
        print("Present =", batch_present_list)
        print("Absent =", batch_absent_list)
        print("Late =", batch_late_list)

    is_filtered = bool(
    course_filter or
    batch_filter or
    student_name or
    status_filter or
    attendance_filter
)
    
    total_attendance = (
    sum(batch_present_list) +
    sum(batch_absent_list) +
    sum(batch_late_list)
)

    show_batch_chart = True
    
    if is_filtered and (course_filter or batch_filter):
        show_batch_chart = total_attendance > 0

    warning_message = ""

    if not show_batch_chart:
        if batch_filter:
            warning_message = (
            "Attendance has not been marked for the selected batch today."
        )

        elif course_filter:
            warning_message = (
            "Attendance has not been marked for the selected course today."
        )
    
    batch_status_map = {}

    today = timezone.now().date()

    for batch in Batch.objects.all():
        marked = Attendance.objects.filter(
        batch=batch,
        attendance_date=today
        ).exists()

        batch_status_map[batch.id] = (
        "not_started" if not marked else "done"
        )

    context = {

        "total_students":
        total_students,

        "present_today":
        present_today,

        "absent_today":
        absent_today,

        "low_attendance":
        low_attendance,

        "report_students":
        report_students,

        "monthly_present":
        monthly_present,

        "monthly_absent":
        monthly_absent,

        "monthly_late":
        monthly_late,

        "course_labels":
        course_labels,

        "course_counts":
        course_counts,

        "batch_labels":
        batch_labels,

        "batch_counts":
        batch_counts,
        
        "batches": batches,
        
        "courses": courses,
        
        "batch_present_counts": batch_present_counts,
        
        "batch_performance_labels": batch_performance_labels,
        
        "batch_performance_counts": batch_performance_counts,
        
        "attendance_status": attendance_status,
        
        "today_marked_count": today_marked_count,
        
        "pending_count": pending_count,
        
        "batch_present_list": batch_present_list,
        
        "batch_absent_list": batch_absent_list,
        
        "batch_percentage_list": batch_percentage_list,
        
        "batch_late_list": batch_late_list,
        
        "batch_chart_title": batch_chart_title,
        
        "monthly_chart_title": monthly_chart_title,
        
        "attendance_date": attendance_date,
        
        "batch_status_map": batch_status_map, 

      
        "is_filtered": is_filtered,
        "show_batch_chart": show_batch_chart,

        "warning_message": warning_message,
     

    }

    return render(
        request,
        "attendance/reports.html",
        context
    )
    
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import enums


def analytics_pdf(request):


    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename=analytics_report.pdf'
    )

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    elements = []

    # =====================
    # TITLE
    # ======================

    elements.append(
        Paragraph(
            "CSC Computer Education",
            styles['Title']
        )
    )

    elements.append(
        Paragraph(
            "Attendance Analytics Report",
            styles['Heading2']
        )
    )

    elements.append(
        Spacer(1, 15)
    )

    # ======================
    # OVERVIEW
    # ======================

    today = timezone.now().date()

    total_students = Enrollment.objects.count()

    present_today = Attendance.objects.filter(
        attendance_date=today,
        status='Present'
    ).count()

    absent_today = Attendance.objects.filter(
        attendance_date=today,
        status='Absent'
    ).count()

    low_attendance = len(
        get_low_attendance_data()
    )

    latest_attendance = Attendance.objects.order_by(
        '-attendance_date'
    ).first()

    attendance_date = "-"

    if latest_attendance:

        attendance_date = (
            latest_attendance.attendance_date.strftime(
                "%d %b %Y"
            )
        )

    summary_data = [

        ["Metric", "Value"],

        ["Total Students", total_students],

        ["Present Today", present_today],

        ["Absent Today", absent_today],

        ["Low Attendance", low_attendance],

        ["Last Attendance Updated", attendance_date]

    ]

    summary_table = Table(
        summary_data,
        colWidths=[220, 150]
    )

    summary_table.setStyle(

        TableStyle([

            ('BACKGROUND',
            (0,0),(-1,0),
            colors.HexColor('#1e40af')),

            ('TEXTCOLOR',
            (0,0),(-1,0),
            colors.white),

            ('FONTNAME',
            (0,0),(-1,0),
            'Helvetica-Bold'),

            ('GRID',
            (0,0),(-1,-1),
            1,
            colors.black),

            ('ROWBACKGROUNDS',
            (0,1),(-1,-1),
            [
                colors.whitesmoke,
                colors.lightgrey
            ])

        ])

    )

    elements.append(summary_table)

    elements.append(
        Spacer(1,20)
    )

    # ======================
    # COURSE & BATCH ANALYTICS
    # ======================

    elements.append(

        Paragraph(
            "Course & Batch Analytics",
            styles['Heading2']
        )

    )

    analytics_data = [[

        "Course",
        "Batch",
        "Students",
        "Total Days",
        "Present",
        "Absent",
        "Late",
        "Attendance %"

    ]]

    for batch in Batch.objects.all():

        students = Enrollment.objects.filter(
            batch=batch
        ).count()

        present = Attendance.objects.filter(
            batch=batch,
            status='Present'
        ).count()

        absent = Attendance.objects.filter(
            batch=batch,
            status='Absent'
        ).count()

        late = Attendance.objects.filter(
            batch=batch,
            status='Late'
        ).count()

        total_days = Attendance.objects.filter(
            batch=batch
        ).count()

        attendance_percentage = round(
            (present / total_days) * 100,
            1
        ) if total_days else 0

        analytics_data.append([

            Paragraph(
                batch.course.course_name,
                styles['BodyText']
            ),

            batch.batch_name,

            students,

            total_days,

            present,

            absent,

            late,

            f"{attendance_percentage}%"

        ])

    analytics_table = Table(

        analytics_data,

        colWidths=[
            140,
            75,
            50,
            55,
            45,
            45,
            40,
            68
        ]

    )

    analytics_table.setStyle(

        TableStyle([

            ('BACKGROUND',
            (0,0),(-1,0),
            colors.HexColor('#2563eb')),

            ('TEXTCOLOR',
            (0,0),(-1,0),
            colors.white),

            ('FONTNAME',
            (0,0),(-1,0),
            'Helvetica-Bold'),

            ('GRID',
            (0,0),(-1,-1),
            1,
            colors.black),

            ('ALIGN',
            (0,0),(-1,-1),
            'CENTER'),

            ('ROWBACKGROUNDS',
            (0,1),(-1,-1),
            [
                colors.whitesmoke,
                colors.lightgrey
            ])

        ])

    )

    elements.append(
        analytics_table
    )

    elements.append(
        Spacer(1,20)
    )

    # ======================
    # ATTENDANCE SUMMARY
    # ======================

    elements.append(

        Paragraph(
            "Attendance Summary",
            styles['Heading2']
        )

    )

    good_count = 0
    warning_count = 0
    critical_count = 0
    excellent_count = 0

    report_students = get_report_students()

    for student in report_students:

        if student["status"] == "Excellent":

            excellent_count += 1
        
        elif student["status"] == "Good":

            good_count += 1

        elif student["status"] == "Warning":

            warning_count += 1

        else:

            critical_count += 1

    attendance_summary = [

        ["Status", "Students"],
        
        ["Excellent", excellent_count],

        ["Good", good_count],

        ["Warning", warning_count],

        ["Critical", critical_count]

    ]

    attendance_summary_table = Table(
        attendance_summary,
        colWidths=[180, 120]
    )

    attendance_summary_table.setStyle(

        TableStyle([

            ('BACKGROUND',
            (0,0),(-1,0),
            colors.HexColor('#059669')),

            ('TEXTCOLOR',
            (0,0),(-1,0),
            colors.white),

            ('FONTNAME',
            (0,0),(-1,0),
            'Helvetica-Bold'),

            ('GRID',
            (0,0),(-1,-1),
            1,
            colors.black),

            ('ALIGN',
            (0,0),(-1,-1),
            'CENTER'),
            
            ('BACKGROUND',
             (0,1),(-1,1),
             colors.green), 
            
            ('BACKGROUND',
            (0,1),(-1,1),
            colors.lightgreen),

            ('BACKGROUND',
            (0,2),(-1,2),
            colors.yellow),

            ('BACKGROUND',
            (0,3),(-1,3),
            colors.salmon),

            ('ROWBACKGROUNDS',
            (0,1),(-1,-1),
            [
                colors.whitesmoke,
                colors.lightgrey
            ])

        ])

    )

    elements.append(
        attendance_summary_table
    )

    doc.build(elements)

    return response



from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from django.http import HttpResponse
from django.utils import timezone


def analytics_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Analytics"

    # ==========================
    # Styles
    # ==========================

    title_fill = PatternFill(
        "solid",
        fgColor="1E3A8A"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="2563EB"
    )

    section_fill = PatternFill(
        "solid",
        fgColor="059669"
    )
    
    excellent_fill = PatternFill(
    "solid",
    fgColor="00B050"
    )

    good_fill = PatternFill(
        "solid",
        fgColor="C6EFCE"
    )

    warning_fill = PatternFill(
        "solid",
        fgColor="FFEB9C"
    )

    critical_fill = PatternFill(
        "solid",
        fgColor="FFC7CE"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==========================
    # Real Time Data
    # ==========================

    today = timezone.now().date()

    total_students = Enrollment.objects.count()

    present_today = Attendance.objects.filter(
        attendance_date=today,
        status='Present'
    ).count()

    absent_today = Attendance.objects.filter(
        attendance_date=today,
        status='Absent'
    ).count()

    low_attendance = len(
        get_low_attendance_data()
    )

    latest_attendance = Attendance.objects.order_by(
        '-attendance_date'
    ).first()

    attendance_date = "-"

    if latest_attendance:

        attendance_date = latest_attendance.attendance_date.strftime(
            "%d %b %Y"
        )

    generated_date = timezone.now().strftime(
        "%d %b %Y %I:%M %p"
    )

    # ==========================
    # Title
    # ==========================

    ws.merge_cells("A1:H1")

    ws["A1"] = (
        "CSC Computer Education - "
        "Attendance Analytics Report"
    )

    ws["A1"].fill = title_fill
    ws["A1"].font = Font(
        color="FFFFFF",
        bold=True,
        size=16
    )
    ws["A1"].alignment = center

    # ==========================
    # Summary
    # ==========================

    ws.merge_cells("A3:B3")

    ws["A3"] = "Summary Dashboard"

    ws["A3"].fill = section_fill
    ws["A3"].font = white_font

    summary_data = [

        ["Metric", "Value"],

        ["Total Students",
         total_students],

        ["Present Today",
         present_today],

        ["Absent Today",
         absent_today],

        ["Low Attendance",
         low_attendance],

        ["Last Attendance Updated",
         attendance_date],

        ["Report Generated On",
         generated_date]

    ]

    row_num = 4

    for row in summary_data:

        ws.append(row)

        if row_num == 4:

            for cell in ws[row_num]:
                cell.fill = header_fill
                cell.font = white_font

        for cell in ws[row_num]:
            cell.border = thin_border

        row_num += 1

    # ==========================
    # Analytics Table
    # ==========================

    start_row = row_num + 2

    ws.merge_cells(
        f"A{start_row}:H{start_row}"
    )

    ws[f"A{start_row}"] = (
        "Course & Batch Analytics"
    )

    ws[f"A{start_row}"].fill = section_fill
    ws[f"A{start_row}"].font = white_font

    analytics_header = [

        "Course",
        "Batch",
        "Students",
        "Total Days",
        "Present",
        "Absent",
        "Late",
        "Attendance %"

    ]

    header_row = start_row + 1

    for col_num, value in enumerate(
        analytics_header,
        start=1
    ):

        cell = ws.cell(
            row=header_row,
            column=col_num
        )

        cell.value = value
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = thin_border

    data_row = header_row + 1

    for batch in Batch.objects.all():

        students = Enrollment.objects.filter(
            batch=batch
        ).count()

        present = Attendance.objects.filter(
            batch=batch,
            status='Present'
        ).count()

        absent = Attendance.objects.filter(
            batch=batch,
            status='Absent'
        ).count()

        late = Attendance.objects.filter(
            batch=batch,
            status='Late'
        ).count()

        total_days = Attendance.objects.filter(
            batch=batch
        ).count()

        attendance_percentage = round(
            (present / total_days) * 100,
            1
        ) if total_days else 0

        row_data = [

            batch.course.course_name,
            batch.batch_name,
            students,
            total_days,
            present,
            absent,
            late,
            f"{attendance_percentage}%"

        ]

        for col_num, value in enumerate(
            row_data,
            start=1
        ):

            cell = ws.cell(
                row=data_row,
                column=col_num
            )

            cell.value = value
            cell.border = thin_border
            cell.alignment = center

        data_row += 1

    # ==========================
    # Attendance Summary
    # ==========================

    good_count = 0
    warning_count = 0
    critical_count = 0
    excellent_count = 0

    total_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()

    for enrollment in Enrollment.objects.all():

        present_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Present'
        ).count()

        attendance_rate = round(
            (present_count / total_days) * 100,
            1
        ) if total_days else 0
        
        if attendance_rate == 100:

            excellent_count += 1

        elif attendance_rate >= 75:

            good_count += 1

        elif attendance_rate >= 60:

            warning_count += 1

        else:

            critical_count += 1

    status_row = data_row + 2

    ws.merge_cells(
        f"A{status_row}:B{status_row}"
    )

    ws[f"A{status_row}"] = (
        "Attendance Status Summary"
    )

    ws[f"A{status_row}"].fill = section_fill
    ws[f"A{status_row}"].font = white_font

    summary_header_row = status_row + 1

    ws.cell(
        summary_header_row,
        1
    ).value = "Status"

    ws.cell(
        summary_header_row,
        2
    ).value = "Students"

    for col in range(1, 3):

        ws.cell(
            summary_header_row,
            col
        ).fill = header_fill

        ws.cell(
            summary_header_row,
            col
        ).font = white_font

    status_data = [
        
        ["Excellent",
         excellent_count,
         excellent_fill],

        ["Good",
         good_count,
         good_fill],

        ["Warning",
         warning_count,
         warning_fill],

        ["Critical",
         critical_count,
         critical_fill]

    ]

    current_row = summary_header_row + 1

    for status, count, color in status_data:

        ws.cell(
            current_row,
            1
        ).value = status

        ws.cell(
            current_row,
            2
        ).value = count

        for col in range(1, 3):

            ws.cell(
                current_row,
                col
            ).fill = color

            ws.cell(
                current_row,
                col
            ).border = thin_border

            ws.cell(
                current_row,
                col
            ).alignment = center

        current_row += 1

    # ==========================
    # Column Width
    # ==========================

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15

    # ==========================
    # Response
    # ==========================

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename=analytics_report.xlsx'
    )

    wb.save(response)

    return response

def report_pdf(request):

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename=attendance_report.pdf'
    )

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    elements = []

    # Title

    elements.append(
        Paragraph(
            "CSC Computer Education",
            styles['Title']
        )
    )

    elements.append(
        Paragraph(
            "Student Attendance Report",
            styles['Heading2']
        )
    )

    latest_attendance = Attendance.objects.order_by(
        '-attendance_date'
    ).first()

    attendance_date = "-"

    if latest_attendance:

        attendance_date = (
            latest_attendance.attendance_date.strftime(
                "%d %b %Y"
            )
        )

    elements.append(

        Paragraph(

            f"Last Attendance Updated : {attendance_date}",

            styles['Normal']

        )

    )

    elements.append(
        Spacer(1, 15)
    )

    # Table Data

    data = [[

        "Student Name",
        "Course",
        "Batch",
        "Present",
        "Absent",
        "Late",
        "Attendance %",
        "Status"

    ]]

    students = get_report_students()

    for student in students:

        data.append([

            f"{student['student'].first_name} "
            f"{student['student'].last_name}",

            student['course'].course_name,

            student['batch'].batch_name
            if student['batch']
            else "-",

            student['present_count'],

            student['absent_count'],

            student['late_count'],

            f"{student['attendance_rate']}%",

            student['status']

        ])

    table = Table(
        data,
        colWidths=[
            90,
            105,
            65,
            45,
            45,
            35,
            75,
            65
        ]
    )

    table_style = TableStyle([

        (
            'BACKGROUND',
            (0, 0),
            (-1, 0),
            colors.HexColor('#1E40AF')
        ),

        (
            'TEXTCOLOR',
            (0, 0),
            (-1, 0),
            colors.white
        ),

        (
            'FONTNAME',
            (0, 0),
            (-1, 0),
            'Helvetica-Bold'
        ),

        (
            'GRID',
            (0, 0),
            (-1, -1),
            1,
            colors.black
        ),

        (
            'ALIGN',
            (0, 0),
            (-1, -1),
            'CENTER'
        ),

        (
            'VALIGN',
            (0, 0),
            (-1, -1),
            'MIDDLE'
        ),

        (
            'ROWBACKGROUNDS',
            (0, 1),
            (-1, -1),
            [
                colors.white,
                colors.HexColor('#F3F4F6')
            ]
        )

    ])

    # Status & Attendance Color

    for row_num, student in enumerate(
        students,
        start=1
    ):

        status = student['status']

        attendance_rate = (
            student['attendance_rate']
        )

        # Status Color

        if status == "Excellent":

            table_style.add(
                'TEXTCOLOR',
                (7, row_num),
                (7, row_num),
                colors.HexColor('#00B050')
            )

        elif status == "Good":

            table_style.add(
                'TEXTCOLOR',
                (7, row_num),
                (7, row_num),
                colors.HexColor('#2563EB')
            )

        elif status == "Warning":

            table_style.add(
                'TEXTCOLOR',
                (7, row_num),
                (7, row_num),
                colors.HexColor('#F59E0B')
            )

        else:

            table_style.add(
                'TEXTCOLOR',
                (7, row_num),
                (7, row_num),
                colors.HexColor('#DC2626')
            )

        # Attendance % Color

        if attendance_rate == 100:

            table_style.add(
                'TEXTCOLOR',
                (6, row_num),
                (6, row_num),
                colors.HexColor('#00B050')
            )

        elif attendance_rate >= 75:

            table_style.add(
                'TEXTCOLOR',
                (6, row_num),
                (6, row_num),
                colors.HexColor('#2563EB')
            )

        elif attendance_rate >= 60:

            table_style.add(
                'TEXTCOLOR',
                (6, row_num),
                (6, row_num),
                colors.HexColor('#F59E0B')
            )

        else:

            table_style.add(
                'TEXTCOLOR',
                (6, row_num),
                (6, row_num),
                colors.HexColor('#DC2626')
            )

    table.setStyle(
        table_style
    )

    elements.append(
        table
    )

    doc.build(
        elements
    )

    return response

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

def report_excel(request):


    wb = Workbook()

    ws = wb.active

    ws.title = "Attendance Report"

    # ---------- TITLE ----------

    ws.merge_cells('A1:H1')

    ws['A1'] = "CSC Computer Education"

    ws['A1'].font = Font(
        size=18,
        bold=True,
        color="1E293B"
    )

    ws['A1'].alignment = Alignment(
        horizontal="center"
    )

    ws.merge_cells('A2:H2')

    ws['A2'] = "Student Attendance Report"

    ws['A2'].font = Font(
        size=14,
        bold=True
    )

    ws['A2'].alignment = Alignment(
        horizontal="center"
    )

    # ---------- DATES ----------

    latest_attendance = Attendance.objects.order_by(
        '-attendance_date'
    ).first()

    attendance_date = "-"

    if latest_attendance:

        attendance_date = (
            latest_attendance.attendance_date.strftime(
                "%d %b %Y"
            )
        )

    ws['A4'] = "Generated On"
    ws['B4'] = timezone.now().strftime(
        "%d %b %Y"
    )

    ws['D4'] = "Last Attendance Updated"
    ws['E4'] = attendance_date

    # ---------- HEADER ----------

    headers = [

        "Student Name",
        "Course",
        "Batch",
        "Present",
        "Absent",
        "Late",
        "Attendance %",
        "Status"

    ]

    header_row = 6

    for col_num, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=header_row,
            column=col_num
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1E40AF"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    students = get_report_students()

    row_num = 7

    thin_border = Border(

        left=Side(style='thin'),

        right=Side(style='thin'),

        top=Side(style='thin'),

        bottom=Side(style='thin')

    )

    excellent_count = 0
    good_count = 0
    warning_count = 0
    critical_count = 0

    for student in students:

        ws.cell(
            row=row_num,
            column=1
        ).value = (
            f"{student['student'].first_name} "
            f"{student['student'].last_name}"
        )

        ws.cell(
            row=row_num,
            column=2
        ).value = (
            student['course'].course_name
        )

        ws.cell(
            row=row_num,
            column=3
        ).value = (
            student['batch'].batch_name
            if student['batch']
            else "-"
        )

        ws.cell(
            row=row_num,
            column=4
        ).value = (
            student['present_count']
        )

        ws.cell(
            row=row_num,
            column=5
        ).value = (
            student['absent_count']
        )

        ws.cell(
            row=row_num,
            column=6
        ).value = (
            student['late_count']
        )

        ws.cell(
            row=row_num,
            column=7
        ).value = (
            f"{student['attendance_rate']}%"
        )

        ws.cell(
            row=row_num,
            column=8
        ).value = (
            student['status']
        )

        # Alternate Row Color

        if row_num % 2 == 0:

            for col in range(1, 9):

                ws.cell(
                    row=row_num,
                    column=col
                ).fill = PatternFill(
                    "solid",
                    fgColor="F8FAFC"
                )

        # Attendance %

        attendance_cell = ws.cell(
            row=row_num,
            column=7
        )

        rate = student[
            'attendance_rate'
        ]

        if rate == 100:

            attendance_cell.font = Font(
                color="00B050",
                bold=True
            )

        elif rate >= 76:

            attendance_cell.font = Font(
                color="2563EB",
                bold=True
            )

        elif rate >= 60:

            attendance_cell.font = Font(
                color="F59E0B",
                bold=True
            )

        else:

            attendance_cell.font = Font(
                color="DC2626",
                bold=True
            )

        # Status Colors

        status_cell = ws.cell(
            row=row_num,
            column=8
        )

        if student['status'] == "Excellent":

            excellent_count += 1

            status_cell.fill = PatternFill(
                "solid",
                fgColor="00B050"
            )

            status_cell.font = Font(
                color="FFFFFF",
                bold=True
            )

        elif student['status'] == "Good":

            good_count += 1

            status_cell.fill = PatternFill(
                "solid",
                fgColor="2563EB"
            )

            status_cell.font = Font(
                color="FFFFFF",
                bold=True
            )

        elif student['status'] == "Warning":

            warning_count += 1

            status_cell.fill = PatternFill(
                "solid",
                fgColor="F59E0B"
            )

            status_cell.font = Font(
                color="FFFFFF",
                bold=True
            )

        else:

            critical_count += 1

            status_cell.fill = PatternFill(
                "solid",
                fgColor="DC2626"
            )

            status_cell.font = Font(
                color="FFFFFF",
                bold=True
            )

        for col in range(1, 9):

            cell = ws.cell(
                row=row_num,
                column=col
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.border = thin_border

        row_num += 1

    # ---------- SUMMARY ----------

    summary_row = row_num + 2

    ws.merge_cells(
        f'A{summary_row}:H{summary_row}'
    )

    ws[
        f'A{summary_row}'
    ] = "Attendance Summary"

    ws[
        f'A{summary_row}'
    ].font = Font(
        bold=True,
        size=14
    )

    ws.cell(
        summary_row + 1,
        1
    ).value = "Excellent"

    ws.cell(
        summary_row + 1,
        2
    ).value = excellent_count

    ws.cell(
        summary_row + 2,
        1
    ).value = "Good"

    ws.cell(
        summary_row + 2,
        2
    ).value = good_count

    ws.cell(
        summary_row + 3,
        1
    ).value = "Warning"

    ws.cell(
        summary_row + 3,
        2
    ).value = warning_count

    ws.cell(
        summary_row + 4,
        1
    ).value = "Critical"

    ws.cell(
        summary_row + 4,
        2
    ).value = critical_count

    # ---------- COLUMN WIDTH ----------

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    response = HttpResponse(

        content_type=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    response['Content-Disposition'] = (

        'attachment; filename=attendance_report.xlsx'

    )

    wb.save(response)

    return response

